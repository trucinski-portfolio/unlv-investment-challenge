#!/usr/bin/env python3
"""
ASTRYX INVESTING - Daily Market Evaluation

NOTE: Prefer using main.py instead:
    python3 main.py scan                    # Full S&P 500 scan → Excel
    python3 main.py scan --watchlist        # Scan watchlist only
    python3 main.py scan --date 2026-01-27  # Historical scan
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import sys
import argparse
from typing import Optional

# Add src to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import (
    WATCHLIST,
    SECTOR_ETFS,
    SCREENING,
    FUNDAMENTAL_LABELS,
)
from data_service import (
    DataService,
    get_sp500_tickers,
    fetch_spy_benchmark,
)
from indicators import add_all_indicators, get_latest_indicators
from screener import StockScreener

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Note: Install openpyxl for Excel support: pip install openpyxl")


def run_daily_evaluation(tickers: list, period: str = "1y",
                         eval_date: Optional[str] = None,
                         fetch_fundamentals: bool = True) -> dict:
    """
    Run complete daily market evaluation.

    Args:
        tickers: List of ticker symbols to scan
        period: Historical data period (e.g., "1y", "6mo")
        eval_date: Optional date string (YYYY-MM-DD) for historical data.
        fetch_fundamentals: Whether to fetch fundamental data (P/E, growth, etc.)

    Returns:
        dict with screening results, setups, sector performance, etc.
    """
    data_service = DataService()

    # Determine the evaluation date
    target_date = None
    if eval_date:
        try:
            target_date = pd.to_datetime(eval_date).date()
            date_display = target_date.strftime('%Y-%m-%d')
            print(f"\n*** HISTORICAL MODE: Getting closing prices for {date_display} ***")
        except ValueError:
            print(f"Invalid date format: {eval_date}. Use YYYY-MM-DD. Falling back to today.")
            eval_date = None
            date_display = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    else:
        date_display = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    print("\n" + "=" * 70)
    print("ASTRYX INVESTING - DAILY MARKET EVALUATION")
    print(f"Date: {date_display}")
    print(f"Stocks: {len(tickers)}")
    print("=" * 70)

    results = {
        'run_date': pd.to_datetime(eval_date) if eval_date else datetime.now(),
        'tickers_scanned': len(tickers),
        'eval_date': eval_date
    }

    # 1. Fetch SPY benchmark
    print("\n[1/5] Fetching benchmark data...")
    spy_df = data_service.fetch_spy(period)
    if spy_df is not None:
        if target_date:
            spy_df = data_service.filter_to_date(spy_df, target_date)
            actual_date = pd.to_datetime(spy_df.index[-1]).date() if not isinstance(spy_df.index[0], str) else pd.to_datetime(spy_df['Date'].iloc[-1]).date()
            if actual_date != target_date:
                print(f"    Note: Using closest available date: {actual_date}")

        spy_df = add_all_indicators(spy_df)
        spy_latest = spy_df.iloc[-1]
        results['spy_price'] = float(spy_latest['Close'])
        results['spy_rsi'] = float(spy_df['RSI'].iloc[-1]) if 'RSI' in spy_df.columns else None

    # 2. Fetch and process all stock data
    print("\n[2/5] Fetching stock data...")
    stock_data = data_service.fetch_multiple(tickers, period=period)

    print("\n[3/5] Calculating indicators...")
    latest_data = {}
    full_data = {}

    for ticker, df in stock_data.items():
        try:
            if target_date:
                df = data_service.filter_to_date(df, target_date)

            df_ind = add_all_indicators(df, spy_df)
            full_data[ticker] = df_ind
            latest_data[ticker] = get_latest_indicators(df_ind)
        except Exception as e:
            print(f"  Error processing {ticker}: {e}")

    print(f"  Processed {len(latest_data)} stocks")

    # 3. Run screener
    print("\n[4/5] Running screener...")
    screener = StockScreener()
    screen_results = screener.screen_multiple(latest_data)
    results['screening_df'] = screen_results

    # 4. Identify setups
    bullish = screen_results[screen_results['Signal'].str.contains('BULLISH', na=False)]
    bearish = screen_results[screen_results['Signal'].str.contains('BEARISH', na=False)]
    neutral = screen_results[screen_results['Signal'] == 'NEUTRAL']

    results['bullish_count'] = len(bullish)
    results['bearish_count'] = len(bearish)
    results['neutral_count'] = len(neutral)
    results['bullish_df'] = bullish
    results['bearish_df'] = bearish

    # 5. Fetch fundamentals
    if fetch_fundamentals:
        print("\n[5/5] Fetching fundamentals...")
        successful_tickers = list(latest_data.keys())
        fundamentals_df = data_service.fetch_fundamentals_batch(successful_tickers)
        results['fundamentals_df'] = fundamentals_df
    else:
        print("\n[5/5] Skipping fundamentals (--no-fundamentals)")
        results['fundamentals_df'] = pd.DataFrame()

    # 6. Sector performance
    print("\nFetching sector ETF data...")
    sector_data = []
    sector_period = '1mo' if target_date else '5d'

    for etf, name in SECTOR_ETFS.items():
        try:
            etf_df = data_service.fetch_stock(etf, period=sector_period, use_cache=False)
            if etf_df is not None and not etf_df.empty:
                if target_date:
                    etf_df = data_service.filter_to_date(etf_df, target_date)

                current = data_service.get_closing_price(etf_df)
                if len(etf_df) > 1:
                    prev = data_service.get_closing_price(etf_df.iloc[:-1])
                else:
                    prev = current

                change_pct = ((current - prev) / prev) * 100 if prev > 0 else 0
                sector_data.append({
                    'ETF': etf,
                    'Sector': name,
                    'Price': current,
                    'Change %': change_pct
                })
        except Exception:
            pass

    results['sectors_df'] = pd.DataFrame(sector_data)

    return results


def create_excel_report(results: dict, output_path: str):
    """
    Create formatted Excel workbook with multiple sheets.

    Sheets:
    1. Summary — market overview
    2. Screening Results — full screening data
    3. Bullish Setups — bullish signals only
    4. Bearish Setups — bearish signals only
    5. Fundamentals — valuation, growth, profitability, health
    6. Sector Performance — ETF changes
    7. Daily Log — historical tracking
    """
    if not EXCEL_AVAILABLE:
        print("Excel export not available. Install openpyxl: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    bullish_fill = PatternFill(start_color='86BA90', end_color='86BA90', fill_type='solid')
    bearish_fill = PatternFill(start_color='E63946', end_color='E63946', fill_type='solid')
    neutral_fill = PatternFill(start_color='A8DADC', end_color='A8DADC', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== Sheet 1: Summary ==========
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["ASTRYX INVESTING - DAILY MARKET EVALUATION"],
        [""],
        ["Run Date:", results['run_date'].strftime('%Y-%m-%d %H:%M:%S')],
        ["Stocks Scanned:", results['tickers_scanned']],
        [""],
        ["MARKET OVERVIEW"],
        ["SPY Price:", f"${results.get('spy_price', 0):.2f}"],
        ["SPY RSI:", f"{results.get('spy_rsi', 0):.1f}" if results.get('spy_rsi') else "N/A"],
        [""],
        ["SIGNAL BREAKDOWN"],
        ["Bullish Setups:", results['bullish_count']],
        ["Bearish Setups:", results['bearish_count']],
        ["Neutral:", results['neutral_count']],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["MARKET OVERVIEW", "SIGNAL BREAKDOWN"]:
                cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25

    # ========== Sheet 2: Full Screening Results ==========
    ws_screen = wb.create_sheet("Screening Results")

    screen_df = results['screening_df'].copy()

    # Merge in key fundamental columns if available
    fund_df = results.get('fundamentals_df', pd.DataFrame())
    if not fund_df.empty and 'ticker' in fund_df.columns:
        fund_cols = fund_df[['ticker', 'sector', 'marketCap', 'trailingPE', 'forwardPE']].copy()
        fund_cols = fund_cols.rename(columns={
            'ticker': 'Ticker', 'sector': 'Sector',
            'marketCap': 'Mkt Cap', 'trailingPE': 'P/E', 'forwardPE': 'Fwd P/E'
        })
        screen_df = screen_df.merge(fund_cols, on='Ticker', how='left')

    headers = list(screen_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_screen.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(screen_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_screen.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            if headers[col_idx - 1] == 'Signal':
                if 'BULLISH' in str(value):
                    cell.fill = bullish_fill
                elif 'BEARISH' in str(value):
                    cell.fill = bearish_fill
                elif value == 'NEUTRAL':
                    cell.fill = neutral_fill

    for col_idx, header in enumerate(headers, 1):
        ws_screen.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, len(header) + 2)

    # ========== Sheet 3: Bullish Setups ==========
    ws_bullish = wb.create_sheet("Bullish Setups")
    _write_signal_sheet(ws_bullish, results['bullish_df'], header_font, bullish_fill, border)

    # ========== Sheet 4: Bearish Setups ==========
    ws_bearish = wb.create_sheet("Bearish Setups")
    _write_signal_sheet(ws_bearish, results['bearish_df'], header_font, bearish_fill, border)

    # ========== Sheet 5: Fundamentals ==========
    ws_fund = wb.create_sheet("Fundamentals")

    if not fund_df.empty:
        # Select and order the most useful columns
        display_cols = ['ticker', 'name', 'sector', 'industry', 'marketCap',
                        'trailingPE', 'forwardPE', 'pegRatio',
                        'priceToSalesTrailing12Months', 'enterpriseToEbitda',
                        'revenueGrowth', 'earningsGrowth', 'earningsQuarterlyGrowth',
                        'returnOnEquity', 'profitMargins', 'operatingMargins', 'grossMargins',
                        'freeCashflow',
                        'debtToEquity', 'currentRatio', 'shortPercentOfFloat', 'shortRatio',
                        'beta', 'dividendYield']
        available_cols = [c for c in display_cols if c in fund_df.columns]
        fund_display = fund_df[available_cols].copy()

        # Use human-readable headers
        readable_headers = []
        for col in available_cols:
            readable_headers.append(FUNDAMENTAL_LABELS.get(col, col.replace('_', ' ').title()))

        for col_idx, header in enumerate(readable_headers, 1):
            cell = ws_fund.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, row in enumerate(fund_display.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                col_name = available_cols[col_idx - 1]

                # Format values
                if col_name == 'marketCap' and pd.notna(value) and value:
                    if value >= 1e12:
                        display_val = f"${value / 1e12:.2f}T"
                    else:
                        display_val = f"${value / 1e9:.1f}B"
                elif col_name == 'freeCashflow' and pd.notna(value) and value:
                    display_val = f"${value / 1e9:.1f}B"
                elif col_name in ['revenueGrowth', 'earningsGrowth', 'earningsQuarterlyGrowth',
                                   'returnOnEquity', 'profitMargins', 'operatingMargins',
                                   'grossMargins', 'dividendYield', 'shortPercentOfFloat']:
                    if pd.notna(value) and value is not None:
                        display_val = f"{value * 100:.1f}%"
                    else:
                        display_val = ""
                elif col_name in ['trailingPE', 'forwardPE', 'pegRatio',
                                   'priceToSalesTrailing12Months', 'enterpriseToEbitda',
                                   'debtToEquity', 'currentRatio', 'shortRatio', 'beta']:
                    if pd.notna(value) and value is not None:
                        display_val = f"{value:.2f}"
                    else:
                        display_val = ""
                else:
                    display_val = value if pd.notna(value) else ""

                cell = ws_fund.cell(row=row_idx, column=col_idx, value=display_val)
                cell.border = border

        for col_idx in range(1, len(available_cols) + 1):
            ws_fund.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
    else:
        ws_fund.cell(row=1, column=1, value="Fundamentals not fetched. Run without --no-fundamentals.")

    # ========== Sheet 6: Sector Performance ==========
    ws_sectors = wb.create_sheet("Sector Performance")

    sector_df = results['sectors_df']
    if not sector_df.empty:
        sector_df = sector_df.sort_values('Change %', ascending=False)

        headers = list(sector_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_sectors.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, row in enumerate(sector_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                if headers[col_idx - 1] == 'Price':
                    value = f"${value:.2f}"
                elif headers[col_idx - 1] == 'Change %':
                    cell = ws_sectors.cell(row=row_idx, column=col_idx, value=f"{value:+.2f}%")
                    if value > 0:
                        cell.fill = bullish_fill
                    elif value < 0:
                        cell.fill = bearish_fill
                    cell.border = border
                    continue

                cell = ws_sectors.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

        for col_idx in range(1, len(headers) + 1):
            ws_sectors.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    # ========== Sheet 7: Daily Log ==========
    ws_log = wb.create_sheet("Daily Log")

    log_headers = ['Date', 'SPY Price', 'SPY RSI', 'Bullish', 'Bearish', 'Neutral', 'Top Setup']
    for col_idx, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    top_setup = results['bullish_df']['Ticker'].iloc[0] if not results['bullish_df'].empty else "None"
    log_data = [
        results['run_date'].strftime('%Y-%m-%d'),
        results.get('spy_price', 0),
        results.get('spy_rsi', 0),
        results['bullish_count'],
        results['bearish_count'],
        results['neutral_count'],
        top_setup
    ]
    for col_idx, value in enumerate(log_data, 1):
        cell = ws_log.cell(row=2, column=col_idx, value=value)
        cell.border = border

    for col_idx in range(1, len(log_headers) + 1):
        ws_log.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    # Save workbook
    wb.save(output_path)
    return output_path


def _write_signal_sheet(ws, df, header_font, fill, border):
    """Helper to write a bullish/bearish setup sheet."""
    if df.empty:
        ws.cell(row=1, column=1, value="No setups found")
        return

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.border = border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border


def print_summary(results: dict):
    """Print summary to console"""
    print("\n" + "=" * 70)
    print("DAILY EVALUATION SUMMARY")
    print("=" * 70)

    print(f"\nSPY: ${results.get('spy_price', 0):.2f} | RSI: {results.get('spy_rsi', 0):.1f}")

    print(f"\nSIGNAL BREAKDOWN:")
    print(f"  Bullish: {results['bullish_count']}")
    print(f"  Bearish: {results['bearish_count']}")
    print(f"  Neutral: {results['neutral_count']}")

    if not results['bullish_df'].empty:
        print(f"\nTOP BULLISH SETUPS:")
        for _, row in results['bullish_df'].head(5).iterrows():
            setups = row.get('Bullish_Setups', row['Signal'])
            print(f"  {row['Ticker']:6} | RSI: {row['RSI']:.1f} | {setups}")

    if not results['bearish_df'].empty:
        print(f"\nTOP BEARISH SETUPS:")
        for _, row in results['bearish_df'].head(5).iterrows():
            setups = row.get('Bearish_Setups', row['Signal'])
            print(f"  {row['Ticker']:6} | RSI: {row['RSI']:.1f} | {setups}")

    if not results['sectors_df'].empty:
        print(f"\nSECTOR PERFORMANCE (Today):")
        sector_sorted = results['sectors_df'].sort_values('Change %', ascending=False)
        for _, row in sector_sorted.head(5).iterrows():
            print(f"  {row['ETF']:5} ({row['Sector']:20}) | {row['Change %']:+.2f}%")

    print("\n" + "=" * 70)


def main():
    parser = argparse.ArgumentParser(
        description='Daily Market Evaluation - ASTRYX INVESTING'
    )
    parser.add_argument('--tickers', '-t', nargs='+', help='Specific tickers to scan')
    parser.add_argument('--top', '-n', type=int, default=None, help='Top N S&P 500 (default: all)')
    parser.add_argument('--watchlist', '-w', action='store_true', help='Scan only your watchlist')
    parser.add_argument('--period', '-p', default='1y', help='Historical data period (default: 1y)')
    parser.add_argument('--output', '-o', default=None, help='Output Excel file path')
    parser.add_argument('--no-excel', action='store_true', help='Save to CSV instead of Excel')
    parser.add_argument('--no-fundamentals', action='store_true', help='Skip fundamentals fetch')
    parser.add_argument('--date', '-d', type=str, default=None,
                        help='Evaluation date (YYYY-MM-DD) for historical data')

    args = parser.parse_args()

    # Determine tickers to scan
    if args.tickers:
        tickers = args.tickers
    elif args.watchlist:
        tickers = WATCHLIST
    else:
        tickers = get_sp500_tickers(args.top)

    # Run evaluation
    results = run_daily_evaluation(
        tickers, period=args.period, eval_date=args.date,
        fetch_fundamentals=not args.no_fundamentals,
    )

    # Print summary
    print_summary(results)

    # Save results
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'output')
    os.makedirs(output_dir, exist_ok=True)

    if args.date:
        timestamp = args.date.replace('-', '')
    else:
        timestamp = datetime.now().strftime('%Y%m%d')

    if args.no_excel or not EXCEL_AVAILABLE:
        # CSV fallback
        screen_path = os.path.join(output_dir, f'screening_{timestamp}.csv')
        results['screening_df'].to_csv(screen_path, index=False)
        print(f"\nScreening results saved to: {screen_path}")
    else:
        excel_path = args.output or os.path.join(output_dir, f'daily_eval_{timestamp}.xlsx')
        create_excel_report(results, excel_path)
        print(f"\nExcel report saved to: {excel_path}")

    return results


if __name__ == "__main__":
    main()
